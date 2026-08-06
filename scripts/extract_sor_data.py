"""
Extract Science of Reading dashboard data from the state's literacy review
workbook (August 2026 version - format may change in future cycles).

Usage:
    python extract_sor_data.py path/to/workbook.xlsx [output_dir]

If output_dir is omitted, defaults to ./data relative to this script.

Row-level scratch cleanup (stray notes rows, a SpEd row that doesn't belong
yet) was done directly in the source file before running this, since that
was a one-off editorial pass, not a repeatable rule.

What this script does handle, because it's mechanical and needs to come out
right every time: filling in missing Lookup Code / EPP Code / Program Type
by joining against overall_scores.json (the canonical EPP roster), splitting
a row into a Traditional and an Alternative row when an EPP runs both
programs, and rescaling the sheet's 1-4 scores onto the app's 0-3 scale
(see src/lib/constants.js).
"""
import sys
import os
import json
import re
import pandas as pd
import openpyxl

if len(sys.argv) < 2:
    print(__doc__)
    sys.exit(1)

SRC = sys.argv[1]
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = sys.argv[2] if len(sys.argv) > 2 else os.path.join(SCRIPT_DIR, "..", "data")
OUT_DIR = os.path.normpath(OUT_DIR)

# Canonical EPP Name -> [(Program Type, Lookup Code, EPP Code), ...], drawn
# from the existing EPP State Review roster. The SoR workbook doesn't
# reliably carry Lookup Code / EPP Code / Program Type itself - some rows
# have them, some have "-", some are blank - so this is what we resolve
# against.
with open(os.path.join(SCRIPT_DIR, "..", "data", "overall_scores.json")) as f:
    overall_scores = json.load(f)

TYPE_ABBR = {"Traditional": "Tra", "Alternative": "Alt"}

# Same exclusion as extract_data.py (2026-08-06, per the state): the
# Alternative program at these two EPPs is dropped from overall_scores.json,
# so this sheet's own "Alt" row for either one has nothing to resolve
# against - expected, not a data error (see resolve_identity below).
EXCLUDED_ALT_NAMES = {"John Brown University", "University of Arkansas-Fayetteville"}

canonical_by_name = {}
for row in overall_scores:
    canonical_by_name.setdefault(row["EPP Name"], []).append(
        (TYPE_ABBR[row["type"]], row["Lookup Code"], row["EPP Code"])
    )

# The SoR sheet spells some names differently than the canonical roster
# (typos, "at" vs "-", casing). Mapped explicitly rather than fuzzy-matched,
# so a name that doesn't match raises loudly instead of silently pairing
# with the wrong school.
NAME_ALIASES = {
    "University of Arkansas at Pine Bluff": "University of Arkansas-Pine Bluff",
    "University of Arkansas at Monticello": "University of Arkansas-Monticello",
    "University of Arkansas at Fort Smith": "University of Arkansas-Fort Smith",
    "Harding Universtiy": "Harding University",
    "ArPEP": "ArPEP/APPEL",
    "REACH University": "Reach University",
    "iTEACH": "iteach Arkansas",
}


def resolve_identity(raw_name, raw_type):
    """Return [(Program Type, Lookup Code, EPP Code), ...] for a raw SoR row.

    One entry if the program type is already known (or the EPP only runs
    one program); two if the row's type is blank/"-" and the EPP runs both
    Traditional and Alternative - the caller duplicates the row per entry.
    """
    canonical_name = NAME_ALIASES.get(raw_name, raw_name)
    options = canonical_by_name.get(canonical_name)
    if not options:
        raise SystemExit(
            f"No match for '{raw_name}' (mapped to '{canonical_name}') in "
            "overall_scores.json. Add an entry to NAME_ALIASES, or confirm "
            "the name against the EPP roster if it's a genuinely new EPP."
        )
    raw_type = raw_type if raw_type in ("Tra", "Alt") else None
    if raw_type is not None:
        matches = [o for o in options if o[0] == raw_type]
        if not matches:
            if raw_type == "Alt" and canonical_name in EXCLUDED_ALT_NAMES:
                return []
            raise SystemExit(
                f"'{canonical_name}' has no {raw_type} program in overall_scores.json."
            )
        return [matches[0]]
    return options


def last_data_row(ws, key_col=1):
    last = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if v is not None and str(v).strip() != "":
            last = r
    return last


def clean_header(h):
    if h is None:
        return None
    return re.sub(r"\s+", " ", str(h)).strip()


wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Sheet1"]

header = [clean_header(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
end = last_data_row(ws)
rows = list(ws.iter_rows(min_row=2, max_row=end, values_only=True))

df = pd.DataFrame(rows, columns=header)
df = df.loc[:, df.columns.notna()]  # drop trailing unlabeled spacer columns

IDENTITY_COLS = ["EPP", "Lookup Code", "Program Type", "EPP Code"]
SCORE_COLS = [c for c in df.columns if c not in IDENTITY_COLS]

out_rows = []
for _, row in df.iterrows():
    canonical_name = NAME_ALIASES.get(row["EPP"], row["EPP"])
    for ptype, lookup_code, epp_code in resolve_identity(row["EPP"], row["Program Type"]):
        out = {
            "EPP Name": canonical_name,
            "Lookup Code": lookup_code,
            "Program Type": ptype,
            "EPP Code": epp_code,
        }
        for c in SCORE_COLS:
            v = pd.to_numeric(row[c], errors="coerce")
            out[c] = (v - 1) if pd.notna(v) else None
        out_rows.append(out)

out_df = pd.DataFrame(out_rows)
records = json.loads(out_df.to_json(orient="records"))

os.makedirs(OUT_DIR, exist_ok=True)
out_path = os.path.join(OUT_DIR, "science_of_reading.json")
with open(out_path, "w") as f:
    json.dump(records, f, indent=2)

print("science_of_reading", out_df.shape)
