"""
Extract EPP State Review dashboard data from the source Excel workbook
into a set of JSON files for the React dashboard.

Usage:
    python extract_data.py path/to/workbook.xlsx [output_dir]

If output_dir is omitted, defaults to ./data relative to this script.

This replaces the old R data-prep scripts (overall_scores.R, standard_data.R,
standard_1_data.R, standard_2_data.R, standard_3_data.R) with equivalent
logic. Whenever Josh sends an updated workbook, re-run this script against
it and commit the regenerated JSON files - do not hand-edit the JSON.

Sheet names in the workbook are assumed stable. If a sheet gets renamed or
restructured (extra header rows, moved columns), this script will need to be
updated to match - it is NOT resilient to arbitrary workbook changes.
"""
import sys
import re
import json
import os
from collections import Counter
import openpyxl
from openpyxl.utils import column_index_from_string
import pandas as pd

if len(sys.argv) < 2:
    print(__doc__)
    sys.exit(1)

SRC = sys.argv[1]
OUT_DIR = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

wb = openpyxl.load_workbook(SRC, data_only=True)


def parse_range(rng):
    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", rng)
    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    return r1, column_index_from_string(c1), r2, column_index_from_string(c2)


def dedupe_names(names):
    name_counts = Counter([n if (n is not None and str(n).strip() != "") else None for n in names])
    result = []
    for i, n in enumerate(names, start=1):
        key = n if (n is not None and str(n).strip() != "") else None
        if key is None:
            result.append(f"...{i}")
        elif name_counts[key] and name_counts[key] > 1:
            result.append(f"{key}...{i}")
        else:
            result.append(key)
    return result


def last_data_row(sheet, key_col=1):
    """Last row in `sheet` with a non-empty value in `key_col` (default col A).

    Row counts grow every year as EPPs are added, so end-of-range must be
    derived from the workbook rather than hardcoded - otherwise new EPPs get
    silently truncated off the bottom, or a smaller workbook yields phantom
    all-null rows. Scoring legends / notes below the table sit in other
    columns, so column A is a reliable signal for where the table ends.
    """
    ws = wb[sheet]
    last = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if v is not None and str(v).strip() != "":
            last = r
    return last


def read_range(sheet, rng, col_names=None):
    ws = wb[sheet]
    r1, c1, r2, c2 = parse_range(rng)
    rows = list(ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2, values_only=True))
    if col_names is None:
        header = dedupe_names(list(rows[0]))
        data = rows[1:]
    else:
        header = col_names
        data = rows
    return pd.DataFrame(data, columns=header)


# ---------------------------------------------------------------
# 1. overall_scores  (Overview page)
# ---------------------------------------------------------------
def build_overall_scores():
    def block(sheet, base_range, s1r, s2r, s3r, ptype):
        base = read_range(sheet, base_range)
        s1 = read_range(sheet, s1r).add_prefix("Standard 1 ")
        s2 = read_range(sheet, s2r).add_prefix("Standard 2 ")
        s3 = read_range(sheet, s3r).add_prefix("Standard 3 ")
        out = pd.concat([base.reset_index(drop=True), s1.reset_index(drop=True),
                          s2.reset_index(drop=True), s3.reset_index(drop=True)], axis=1)
        out["type"] = ptype
        return out

    t_end = last_data_row("Traditional EPP Score Totals")
    a_end = last_data_row("Alternative EPP Score Total")
    trad = block("Traditional EPP Score Totals",
                 f"A3:E{t_end}", f"N3:P{t_end}", f"AC3:AE{t_end}", f"AS3:AU{t_end}", "Traditional")
    alt = block("Alternative EPP Score Total",
                f"A3:E{a_end}", f"N3:P{a_end}", f"AC3:AE{a_end}", f"AS3:AU{a_end}", "Alternative")
    overall = pd.concat([trad, alt], ignore_index=True)
    return overall


# ---------------------------------------------------------------
# 2. standard_data  (Standard 1/2/3 pages, Summary view)
# ---------------------------------------------------------------
def build_standard_data():
    s1_cols = ["1.1(a)", "1.1", "1.2(a)", "1.2(b)", "1.2", "1.3(a)", "1.3",
               "Standard 1 Average Criteria Score", "Standard 1 Performance Level", "Standard 1 Performance Score"]
    s2_cols = ["2.1(a)", "2.1(b)", "2.1(c)", "2.1", "2.2(a)", "2.2(b)", "2.2(c)", "2.2",
               "2.3(a)", "2.3(b)", "2.3", "Standard 2 Average Criteria Score",
               "Standard 2 Performance Level", "Standard 2 Performance Score"]
    s3_cols = ["3.1(a)", "3.1(b)", "3.1(c)", "3.1", "3.2(a)", "3.2(b)", "3.2(c)", "3.2",
               "3.3(a)", "3.3(b)", "3.3(c)", "3.3", "Standard 3 Average Criteria Score",
               "Standard 3 Performance Level", "Standard 3 Performance Score"]

    def block(sheet, names_range, s1_range, s2_range, s3_range, ptype):
        names = read_range(sheet, names_range)
        s1 = read_range(sheet, s1_range, col_names=s1_cols)
        s2 = read_range(sheet, s2_range, col_names=s2_cols)
        s3 = read_range(sheet, s3_range, col_names=s3_cols)
        for c in s3.columns:
            if c.startswith("3"):
                s3[c] = pd.to_numeric(s3[c], errors="coerce").round(2)
        out = pd.concat([names.reset_index(drop=True), s1.reset_index(drop=True),
                          s2.reset_index(drop=True), s3.reset_index(drop=True)], axis=1)
        out["type"] = ptype
        return out

    t_end = last_data_row("Traditional EPP Score Totals")
    a_end = last_data_row("Alternative EPP Score Total")
    trad = block("Traditional EPP Score Totals",
                 f"A3:C{t_end}", f"G4:P{t_end}", f"R4:AE{t_end}", f"AG4:AU{t_end}", "Traditional")
    alt = block("Alternative EPP Score Total",
                f"A3:C{a_end}", f"G4:P{a_end}", f"R4:AE{a_end}", f"AG4:AU{a_end}", "Alternative")
    return pd.concat([trad, alt], ignore_index=True)


# ---------------------------------------------------------------
# 3. standard_1_data  (Standard 1 Detail view)
# ---------------------------------------------------------------
def build_standard_1_data():
    ws = wb["1.2-1.3 Enrollment & Completion"]
    header = [ws.cell(row=2, column=c).value for c in range(1, 30)]

    # The 6 "New Enrollees" year columns (positions 5-10) roll forward every
    # year - read the real years off the header dynamically rather than
    # hardcoding them, so this keeps working no matter what year it is.
    year_re = re.compile(r"(\d{4})")
    enroll_years = []
    for h in header[4:10]:
        m = year_re.match(str(h))
        if not m:
            raise SystemExit(
                "Could not read a year from the enrollment column headers on "
                "'1.2-1.3 Enrollment & Completion' (row 2, columns E-J). "
                f"Got: {header[4:10]!r}. The sheet layout has probably changed - "
                "check the workbook before trusting any output."
            )
        enroll_years.append(m.group(1))
    prior_years = enroll_years[0:3]
    recent_years = enroll_years[3:6]

    generic_cols = [
        "EPP Name", "Lookup Code", "Program Type", "EPP Code",
        "enroll_1", "enroll_2", "enroll_3", "enroll_4", "enroll_5", "enroll_6",
        "completers_1", "completers_2", "completers_3",
        "enroll_subj_4", "enroll_subj_5", "enroll_subj_6",
        "shortage_4", "shortage_5", "shortage_6",
        "_spacer1",
        "Average 3yr Growth", "score_avg_3_year_growth",
        "_spacer2",
        "Percentage of New Enrollees in Subject Shortage Areas", "score_new_in_shortage",
        "_spacer3",
        "3yr Completion Rate", "score_3_year_completion",
        "_spacer4",
    ]
    # data starts at row 3 (row 2 is the header we just read manually above)
    s1_end = last_data_row("1.2-1.3 Enrollment & Completion")
    raw = read_range("1.2-1.3 Enrollment & Completion", f"A3:AC{s1_end}", col_names=generic_cols)

    numeric_cols = ["enroll_1", "enroll_2", "enroll_3", "enroll_4", "enroll_5", "enroll_6",
                    "completers_1", "completers_2", "completers_3",
                    "enroll_subj_4", "enroll_subj_5", "enroll_subj_6",
                    "shortage_4", "shortage_5", "shortage_6"]
    for col in numeric_cols:
        raw[col] = pd.to_numeric(raw[col], errors="coerce")

    raw["new_enrollees_prior_3yr"] = raw["enroll_1"] + raw["enroll_2"] + raw["enroll_3"]
    raw["new_enrollees_recent_3yr"] = raw["enroll_4"] + raw["enroll_5"] + raw["enroll_6"]
    raw["completers_prior_3yr"] = raw["completers_1"] + raw["completers_2"] + raw["completers_3"]
    raw["new_enrollees_w_subject_recent_3yr"] = raw["enroll_subj_4"] + raw["enroll_subj_5"] + raw["enroll_subj_6"]
    raw["shortage_enrollment_recent_3yr"] = raw["shortage_4"] + raw["shortage_5"] + raw["shortage_6"]

    out = raw[[
        "EPP Name", "Lookup Code", "Program Type", "EPP Code",
        "new_enrollees_prior_3yr", "new_enrollees_recent_3yr", "completers_prior_3yr",
        "new_enrollees_w_subject_recent_3yr", "shortage_enrollment_recent_3yr",
        "Average 3yr Growth", "score_avg_3_year_growth",
        "Percentage of New Enrollees in Subject Shortage Areas", "score_new_in_shortage",
        "3yr Completion Rate", "score_3_year_completion",
    ]].copy()

    year_labels = {
        "new_enrollees_prior_3yr_years": f"{prior_years[0]}-{prior_years[-1]}",
        "new_enrollees_recent_3yr_years": f"{recent_years[0]}-{recent_years[-1]}",
        "completers_prior_3yr_years": f"{prior_years[0]}-{prior_years[-1]}",
        "new_enrollees_w_subject_recent_3yr_years": f"{recent_years[0]}-{recent_years[-1]}",
        "shortage_enrollment_recent_3yr_years": f"{recent_years[0]}-{recent_years[-1]}",
    }
    # denormalize onto every row rather than shipping a second JSON file -
    # cheap at this row count, and means the frontend never has to join two
    # files together just to label a column
    for k, v in year_labels.items():
        out[k] = v

    return out


# ---------------------------------------------------------------
# 4. standard_2_data  (Standard 2 Detail view) - whole sheet as-is
# ---------------------------------------------------------------
def build_standard_2_data():
    ws = wb["2.1-2.3 Preparing Candidates"]
    rows = list(ws.iter_rows(values_only=True))
    header = dedupe_names(list(rows[0]))
    df = pd.DataFrame(rows[1:], columns=header)
    # drop fully-empty trailing columns (blank spacer columns with no header and no data)
    df = df.dropna(axis=1, how="all")
    return df


# ---------------------------------------------------------------
# 5. standard_3_data  (Standard 3 Detail view) - 3.1, 3.2, 3.3 joined
# ---------------------------------------------------------------
def build_standard_3_data():
    join_keys = ["EPP", "Lookup Code", "Program Type", "EPP Code"]

    a = read_range("3.1(a) - Standard License", f"A1:L{last_data_row('3.1(a) - Standard License')}")
    b = read_range("3.1(b) - Provisional License", f"A1:L{last_data_row('3.1(b) - Provisional License')}")
    c = read_range("3.1(c) - Praxis Pass Rates", f"A1:K{last_data_row('3.1(c) - Praxis Pass Rates')}")
    c["Program Type"] = c.apply(lambda row: "Tra" if row["EPP"] != "Totals" else None, axis=1)

    s31 = a.merge(b, on=join_keys, how="outer", suffixes=("_3_1_a", "_3_1_b"))
    s31 = s31.merge(c, on=join_keys, how="outer", suffixes=("", "_3_1_c"))

    s32 = read_range("3.2(a)-(c) - Employment", f"A2:Y{last_data_row('3.2(a)-(c) - Employment')}")
    s32 = s32.rename(columns={
        "...3": "Program Type",
        "Score...19": "score_employed",
        "Score...22": "score_priority_employment",
        "Score...25": "score_experience",
    })
    s32 = s32.drop(columns=["...17", "...20", "...23"], errors="ignore")

    a3 = read_range("3.3(a) - Supervisor Survey", f"A1:H{last_data_row('3.3(a) - Supervisor Survey')}")
    b3 = read_range("3.3(b) - Completer Survey", f"A1:H{last_data_row('3.3(b) - Completer Survey')}")
    c3 = read_range("3.3(c) - Effectiveness", f"A1:L{last_data_row('3.3(c) - Effectiveness')}")
    s33 = a3.merge(b3, on=join_keys, how="outer", suffixes=("_3_3_a", "_3_3_b"))
    s33 = s33.merge(c3, on=join_keys, how="outer", suffixes=("", "_3_3_c"))

    return s31, s32, s33


def to_records(df):
    # Replace NaN with None, keep everything else as-is
    return json.loads(df.to_json(orient="records", date_format="iso"))


# Performance Level thresholds (Exceeds/Meets/Approaching/Below) are fixed
# business rules, not derived from the workbook - they belong hardcoded as
# a constant in the app itself (see brief.md), not generated here.


if __name__ == "__main__":
    out_dir = OUT_DIR
    os.makedirs(out_dir, exist_ok=True)

    overall_scores = build_overall_scores()
    standard_data = build_standard_data()
    standard_1_data = build_standard_1_data()
    standard_2_data = build_standard_2_data()
    s31, s32, s33 = build_standard_3_data()

    datasets = {
        "overall_scores": overall_scores,
        "standard_data": standard_data,
        "standard_1_data": standard_1_data,
        "standard_2_data": standard_2_data,
        "standard_3_data_3_1": s31,
        "standard_3_data_3_2": s32,
        "standard_3_data_3_3": s33,
    }

    for name, df in datasets.items():
        with open(f"{out_dir}/{name}.json", "w") as f:
            json.dump(to_records(df), f, indent=2)
        print(name, df.shape)
